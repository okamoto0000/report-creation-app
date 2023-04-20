import pandas as pd
import datetime
from datetime import timedelta
import pandas.tseries.offsets as offsets
import matplotlib.pyplot as plt
import sklearn
from sklearn.linear_model import LinearRegression


import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image



#▲レポート作成
#日付の変数設定
today = datetime.date(2023,9,14)
month_start = datetime.date(today.year, today.month, 1)
month_end = datetime.date(today.year, today.month+1, 1) - timedelta(days=1)

#変数設定、Excel読み込み
import_file = "image_file_storage/send_file.xlsx"
wb = openpyxl.load_workbook("image_file_storage/send_file.xlsx")
excel_sheetname01 = wb.sheetnames[0]
excel_sheetname02 = wb.sheetnames[1]
df_last_year_excel = pd.read_excel(import_file, sheet_name = excel_sheetname01,index_col='date')



#前年比の作成
df_last_year_excel['next_year'] = df_last_year_excel.index + offsets.Day(364)
df_last_year_excel = df_last_year_excel.drop(columns=['cost'],axis=1)
df_this_year_excel = pd.read_excel(import_file, sheet_name = excel_sheetname02,index_col='date')
#去年のデータと今年のデータを結合
df_this_year = pd.merge(df_last_year_excel,df_this_year_excel,left_on='next_year',right_index=True,how='right')
df_this_year = df_this_year.rename(columns={'sales_y':'this year sales','sales_x':'last year sales','next_year':'date'})
#インデックス変更
df_this_year = df_this_year.set_index('date')
df_this_year = df_this_year.sort_index()
df_this_year['last year per'] = df_this_year['this year sales']/df_this_year['last year sales']



#前週比を計算
df_this_year['last week sales'] = df_this_year['this year sales'].shift(7)
df_this_year['last week per'] = df_this_year['this year sales']/df_this_year['last week sales']
#日付で条件抽出
df_this_year = df_this_year[(df_this_year.index >= pd.to_datetime(month_start)) & (df_this_year.index < pd.to_datetime(today))]
#変数設定(指定した行名)
specified_date = datetime.date(today.year, today.month, today.day-1)
specified_date = specified_date.strftime("%Y-%m-%d")



#棒グラフ作成
#グラフサイズ指定
plt.figure(figsize=(10,7))
            
#X軸方向のデータ配置・位置
x1 = [1, 2]
#棒グラフの高さ・データの取得・棒グラフの太さ0.2
y1 = [df_this_year.loc[specified_date,'this year sales'],df_this_year.loc[specified_date,'this year sales']]

x2 = [1.2]
y2 = [df_this_year.loc[specified_date,'last year sales']]

x3 = [2.2]
y3 = [df_this_year.loc[specified_date,'last week sales']]

#棒グラフの設定(色、名前、太さ、整列)
plt.bar(x1, y1, color='#F7D238',label='this year', width=0.2, align="center")
plt.bar(x2, y2, color='#295C82', label='last year', width=0.2, align="center")
plt.bar(x3, y3, color='#6D9ED8', label='last week', width=0.2, align="center")

#X軸ラベル設定、X軸の名前の位置・設定
label_x = ['this year/last year', 'this week/last week']
plt.xticks([1.15, 2.15], label_x)

#凡例設定
plt.legend(fontsize=25,loc='upper center')
plt.savefig("image_file_storage/graph01.png",dpi=60)



#線グラフ作成
#グラフサイズの設定
plt.figure(figsize=(10,7))

#XY軸の設定
x = df_this_year.index
y1 = df_this_year['this year sales']
y2 = df_this_year['last year sales']
y3 = df_this_year['last week sales']

#線グラフの設定
plt.plot(x, y1, color="#F7D238", label='this year' , linewidth="5")
plt.plot(x, y2, color="#295C82", label='last year')
plt.plot(x, y3, color="#6D9ED8", label='last week')

#凡例設定
plt.legend(fontsize=25,loc='upper left')
plt.savefig("image_file_storage/graph02.png",dpi=60)



#▲回帰分析の設定
df_analysis_last_year = pd.read_excel(import_file, sheet_name = excel_sheetname01,index_col='date')
df_analysis_last_year = df_analysis_last_year[['sales','cost']]

df_analysis_this_year = pd.read_excel(import_file, sheet_name = excel_sheetname02,index_col='date')
df_analysis_this_year = df_analysis_this_year[['sales','cost']]

df_concat = pd.concat([df_analysis_last_year,df_analysis_this_year])

#曜日のカラム作成
df_concat['weekday_name'] = df_concat.index.day_name()
df_concat = df_concat.dropna()

#平日と週末カラム作成
def find_weekend(weekday_name):
    if  (weekday_name == "Saturday") or (weekday_name == "Sunday"):
        return "weekend"
    else:
        return "weekday"
    
df_concat["weekend"] = df_concat["weekday_name"].apply(find_weekend)
df_concat = df_concat.drop(columns='weekday_name',axis=1)
#ダミー変数の作成 0か1に変換
df_dummies = pd.get_dummies(df_concat)

#予測に使うデータ、予測したいデータ作成
x = df_dummies.drop('sales',axis=1)
y = df_dummies['sales']



#▲回帰分析の実行
model = LinearRegression()
model.fit(x,y)

#決定係数の変数設定
accuracy = float('{:.5f}'.format(model.score(x, y)))
if accuracy >= 0.9:
    accuracy = f"精度が高い({accuracy})"
elif accuracy >= 0.8:
    accuracy = f"それなりの精度({accuracy})"
else:
    accuracy = f"精度が低い({accuracy})"

#予測データの作成
date_range = pd.date_range(start=today, end=month_end, freq='D')
df_prediction = pd.DataFrame(index=date_range, columns=['this year sales','this year sales(cost)'])

#平日と週末カラム作成
df_prediction['weekday_name'] = df_prediction.index.day_name()
df_prediction["weekend"] = df_prediction["weekday_name"].apply(find_weekend)

#係数、切片の変数設定
y = model.intercept_ 
x_cost = model.coef_[0]
x_weekday = model.coef_[1]
x_weekend = model.coef_[2]


#広告あり・なしの予測値を算出
cost=10000
for index_name in df_prediction.index:
    if df_prediction.loc[index_name,'weekend'] == "weekend":
        df_prediction.loc[index_name,'this year sales'] = y + x_weekend + x_cost*0
        df_prediction.loc[index_name,'this year sales(cost)'] =  y + x_weekend + x_cost*cost
    else:                                                          
        df_prediction.loc[index_name,'this year sales'] = y + x_weekday + cost*0
        df_prediction.loc[index_name,'this year sales(cost)'] = y + x_weekday + x_cost*cost

df_prediction = df_prediction.drop(columns=['weekday_name','weekend'])
df_this_year['this year sales(cost)']=df_this_year['this year sales']

#実績値と予測値の結合
df_this_year = pd.concat([df_this_year,df_prediction],sort = False)
df_this_year = df_this_year[['this year sales','this year sales(cost)','cost','last week per','last week sales','last year per','last year sales']]

#広告費のあり・なしの売上合計予測の算出
prediction = int(df_this_year['this year sales'].sum())
prediction_cost = int(df_this_year['this year sales(cost)'].sum())





#▲Excelファイルの編集
export_file = "image_file_storage/sales_forecast_report.xlsx"

#Excelファイルの書き出す
df_this_year.to_excel(export_file)
workbook = openpyxl.load_workbook(export_file)
worksheet = workbook.worksheets[0]

#フォント設定・変更
font = Font(name="メイリオ",size=14)
sheet_range = worksheet["A1":"H32"]

for row in sheet_range:
    for cell in row:
        worksheet[cell.coordinate].font = font


#セルの塗りつぶし、幅の調整、フォント変更
fill = openpyxl.styles.PatternFill(patternType="solid",fgColor="295C82",bgColor="295C82")
for col in ["A1","B1","C1","D1","E1","F1","G1","H1"]:
    worksheet[col].fill = fill

for col in ["A","B","D","E","F","G","H"]:
    worksheet.column_dimensions[col].width = 18

worksheet.column_dimensions["C"].width = 24
worksheet.column_dimensions["J"].width = 30
worksheet.column_dimensions["K"].width = 20

for col in ["B1","C1","D1","E1","F1","G1","H1"]:
    worksheet[col].font = Font(name="メイリオ",size=14,color="FFFFFF")


#表示フォーマット指定
for idx in range(2, 33):
    worksheet.cell(row=idx,column=1).number_format = 'yyyy-mm-dd' 
    worksheet.cell(row=idx,column=5).number_format = "0%" 
    worksheet.cell(row=idx,column=7).number_format = "0%"
    worksheet.cell(row=idx,column=2).number_format = "#,##0" 
    worksheet.cell(row=idx,column=3).number_format = "#,##0" 
    worksheet.cell(row=idx,column=4).number_format = "#,##0" 
    worksheet.cell(row=idx,column=6).number_format = "#,##0" 
    worksheet.cell(row=idx,column=8).number_format = "#,##0" 
    worksheet.cell(row=idx,column=11).number_format = "#,##0" 
    worksheet.cell(row=idx,column=1).alignment = Alignment(horizontal='center')


#セルに広告の値とモデル精度の値を付ける・画像を添付
worksheet["J2"].value = "今月着地(広告なし)"
worksheet["J3"].value = "今月着地(広告1万円)"
worksheet["J4"].value = "決定係数(モデル精度)"
worksheet["K2"].value = prediction
worksheet["K3"].value = prediction_cost
worksheet["K4"].value = accuracy
#セルの塗りつぶし
fill = openpyxl.styles.PatternFill(patternType="solid",fgColor="6D9ED8",bgColor="6D9ED8")

for col in ["J2","J3","J4","K2","K3","K4"]:
    worksheet[col].fill = fill

#画像を添付
img1 = Image("image_file_storage/graph01.png")
worksheet.add_image(img1,"I6")
img2 = Image("image_file_storage/graph02.png")
worksheet.add_image(img2,"I20")

workbook.save(export_file)

