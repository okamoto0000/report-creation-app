import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image 


#日付操作、Excel読み込み
import_file = "image_file_storage/send_file.xlsx"
wb = openpyxl.load_workbook("image_file_storage/send_file.xlsx")
excel_sheetname01 = wb.sheetnames[0]
excel_sheetname02 = wb.sheetnames[1]


#Excelデータ読み込み
df_last_year = pd.read_excel(import_file, sheet_name = excel_sheetname01,index_col='date')
df_last_year = df_last_year[['sales','cost']]

df_this_year = pd.read_excel(import_file, sheet_name = excel_sheetname02,index_col='date')
df_this_year = df_this_year[['sales','cost']]


#上下に結合、曜日の追加
df_concat = pd.concat([df_last_year,df_this_year])
df_concat['weekday_name'] = df_concat.index.day_name()
df_concat = df_concat.dropna()



#グループごとの集計
df_last_year_weekdayname = df_concat[['weekday_name','sales']].groupby('weekday_name').max()
df_last_year_weekdayname["min_sales"] = df_concat[['weekday_name','sales']].groupby('weekday_name').min()
df_last_year_weekdayname['sum_sales'] = df_concat[['weekday_name','sales']].groupby('weekday_name').sum()
df_last_year_weekdayname["mean_sales"] = df_concat[['weekday_name','sales']].groupby('weekday_name').mean()
#列名変更
df_last_year_weekdayname = df_last_year_weekdayname.rename(columns={'sales':'max_sales'})



#棒グラフの作成
plt.figure(figsize=(10,7))
plt.bar(df_last_year_weekdayname.index,df_last_year_weekdayname['mean_sales'])
plt.savefig("image_file_storage/graph001.png",dpi=60)

#散布図と回帰直線の作成
plt.figure(figsize=(10,7))
sns.regplot(x=df_concat["cost"], y=df_concat["sales"], data=df_concat)
plt.savefig("image_file_storage/graph002.png",dpi=60)


#▲Excelファイルの編集
export_file = "image_file_storage/sales_forecast_report.xlsx"
#Excelファイルの書き出す
df_last_year_weekdayname.to_excel(export_file)
workbook = openpyxl.load_workbook(export_file)
worksheet = workbook.worksheets[0]


#行・列の移動
worksheet.move_range("A1:E8", rows=26, cols=4, translate=True)
#フォント設定・変更
font = Font(name='メイリオ', size=14)
sheet_range = worksheet['E27':'I34']
for row in sheet_range:
    for cell in row:
        print(cell)
        worksheet[cell.coordinate].font = font



#セルの塗りつぶし、幅の調整、フォント変更
fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='295C82', bgColor='295C82')
for col in ['E27', 'F27', 'G27', 'H27', 'I27']:
    worksheet[col].fill = fill

for col in ['F','G','H','I']:
    worksheet.column_dimensions[col].width = 18
worksheet.column_dimensions['E'].width = 24

for col in ['E27', 'F27', 'G27', 'H27','I27']:
    worksheet[col].font = Font(name='メイリオ', size=14, color="FFFFFF")



#表示フォーマット指定
for idx in range(26, 35):
    worksheet.cell(row=idx,column=6).number_format = "#,##0" 
    worksheet.cell(row=idx,column=7).number_format = "#,##0" 
    worksheet.cell(row=idx,column=8).number_format = "#,##0" 
    worksheet.cell(row=idx,column=9).number_format = "#,##0" 



#画像を添付
img1 = Image('image_file_storage/graph001.png') 
worksheet.add_image(img1, 'B2') 
img2 = Image('image_file_storage/graph002.png') 
worksheet.add_image(img2, 'H2') 

workbook.save(export_file)